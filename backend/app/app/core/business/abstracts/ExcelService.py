from typing import Any, List, Dict
from multidict import MultiDict
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Series, Reference


class ExcelService:
    def __init__(self):
        self.wb = Workbook()

    def writeToExcelFile(self, data: Any, file_name: str, title: str):
        work_sheet = self.wb.active
        work_sheet.title = title
        for row in range(1, 40):
            work_sheet.append(range(600))
        self.save(filename=file_name)


    def readExcelFile(self, filename: str, data:str):
        self.wb = load_workbook(filename=filename)
        sheet_ranges = self.wb[data.get('index').value]


